from django.shortcuts import render_to_response, redirect, render
from django.http import HttpResponseRedirect, HttpResponse
from django.core.files import File
from django.core.exceptions import ValidationError
from django.core.context_processors import csrf 
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.contrib import messages
from django.contrib.auth import authenticate, login,logout
from django.template import Context
from django.template.response import TemplateResponse
from django.template import RequestContext
from django.utils.translation import ugettext, ugettext_lazy as _
from django.utils.encoding import smart_unicode, force_unicode
from django.utils import simplejson
from django.conf import settings

from events.models import *
from events.forms import *
from college_event.models import *
from reviews.models import *
from banner.models import *
from payu.models import *
from events.util import format_redirect_url
from templated_email import send_templated_mail
from forms import UploadFileForm
from django.utils.timezone import utc
import simplejson as json
import random
import string
import datetime
import time
import openpyxl


class JSONResponse(HttpResponse):
	def __init__(self, data):
		super(JSONResponse, self).__init__(
				simplejson.dumps(data), mimetype='application/json')

def home(request):
    if request.user.is_superuser:
        logout(request)
        return HttpResponseRedirect('/')
    return render_to_response("index_v2.html", context_instance=RequestContext(request))

def about(request):
	return render_to_response("about-us.html", context_instance=RequestContext(request))

def privacy(request):
	return render_to_response("privacy.html", context_instance=RequestContext(request))

def terms_and_conditions(request):
	return render_to_response("terms_and_conditions.html", context_instance=RequestContext(request))

def faqs(request):
	return render_to_response("faqs.html", context_instance=RequestContext(request))

def start(request):
	return render_to_response('index_v2.html',context_instance=RequestContext(request))

def logout_view(request):
	logout(request)
	response = HttpResponseRedirect("/")
	return response

def details(request,id=None):
    # try:
    postevent=Postevent.objects.get(pk=id)
    img=str(postevent.poster).split(',')
    photo=img[0]
    photos=[n for n in str(postevent.poster).split(',')]
    organizer=Organizer.objects.filter(postevent__id=postevent.id)
    review=Review.objects.filter(event_id=postevent.id)
    related_events = Postevent.objects.filter(category = postevent.category, eventtype=postevent.eventtype, city=postevent.city)
    return render_to_response("company-profile.html",{'events':postevent,'organizer':organizer,'review':review,'related_events':related_events,'photos':photos,'photo':photo}, context_instance=RequestContext(request))
    # except:
    #     return render_to_response("company-profile.html",{'message':'Sorry for inconvenience.Some thing went to wrong'}, context_instance=RequestContext(request))

def banner(request):
	return render_to_response("uploadbanner.html",context_instance=RequestContext(request))

# old login code when not using ajax
# @csrf_protect 
# def user_login(request):
#     """
#     Login User
#     """
#     logout(request)
#     username = password = ''
#     if request.POST.get("next") is None:
#         return HttpResponseRedirect('/')
#     elif request.POST.get("next"):
#         email = request.POST['username']
#         print 'login email', email
#         password = request.POST['password']
#         user = authenticate(email=email, password=password)
#         print 'user login', user
#         if user is not None:
#             if user.is_active:
#                 login(request, user)
#                 return HttpResponseRedirect(request.POST.get("next"))
#         try:
#             error={}
#             if not User.objects.filter(email=email).exists():
#                 error['email_exists'] = ugettext('Email Does not exists')
#                 raise ValidationError(error['email_exists'], 1)
#             # else:
#             #     if not User.objects.filter(username=username).exists():
#             #         error['username_exists'] = ugettext('Username Does not exists')
#             #         raise ValidationError(error['username_exists'], 2)
#         except ValidationError as e:
#             messages.add_message(request, messages.ERROR, e.messages[-1]) 
#             redirect_path = request.POST["next"]
#             query_string = 'lst=%d' % e.code
#             redirect_url = format_redirect_url(redirect_path, query_string)
#             return HttpResponseRedirect(redirect_url)
#         if not error:
#             # if not '@' in username:
#             #     user = User.objects.get(username=username)
#             # else:
#             user = User.objects.get(email=email)
#             user.backend='django.contrib.auth.backends.ModelBackend'
#             try:
#                 error={}
#                 if user.check_password(password):
#                     print user
#                 else:
#                     error['password'] = ugettext('Invalid password')
#                     raise ValidationError(error['password'], 3)
#             except ValidationError as e:
#                 messages.add_message(request, messages.ERROR, e.messages[-1]) 
#                 redirect_path = request.POST["next"]
#                 query_string = 'lst=%d' % e.code
#                 redirect_url = format_redirect_url(redirect_path, query_string)
#                 return HttpResponseRedirect(redirect_url) 
#             if user:           
#                 if user.is_active:                
#                     login(request, user)
#                     user_id=user.id
#                     response=HttpResponseRedirect(request.POST["next"]) 
#                     return response               
#     else:
#         email = request.POST['username']
#         password = request.POST['password']
#         user = authenticate(email=email, password=password)
#         print 'user login else', user
#         if user is not None:
#             if user.is_active:
#                 login(request, user)
#                 return HttpResponseRedirect('/')

# new login code when using ajax (updated by kalai)
@csrf_exempt
def user_login(request):    
    import json 
    if request.user.is_superuser:
        logout(request)
        return HttpResponseRedirect('/')        
    logout(request)
    error = {}
    username = request.POST['username']
    print "username", username
    password = request.POST['password']
    print "password", password
    context = {}
    if not User.objects.filter(email=username).exists():
        error['email_exists'] = True
        response = HttpResponse(json.dumps(error, ensure_ascii=False),mimetype='application/json')
    else:
        user = User.objects.get(email=username)
        user.backend='django.contrib.auth.backends.ModelBackend'
        if user:
            if user.check_password(password):
                if user.is_active:
                    login(request, user)
                    response = HttpResponseRedirect(request.POST.get('next')) 
            else:
                error['password'] = True
                response = HttpResponse(json.dumps(error, ensure_ascii=False),mimetype='application/json')           
    return response

@csrf_protect
def register(request): 
	print "register" 
	context = RequestContext(request) 
	registered = False
	user=User()
	userprofile=Userprofile()
	if request.method == 'POST': 
		email=request.POST['email_id']      
		username=request.POST['username']       
		try:
			error={}
			# if User.objects.filter(username=username).exists():
			#     error['username_exists'] = ugettext('Username already exists')
			#     raise ValidationError(error['username_exists'], 1)
			if User.objects.filter(email=email).exists():
				print 'User.objects.filter(email=email).exists()', User.objects.filter(email=email).exists()
				error['email_exists'] = ugettext('Email already exists')
				raise ValidationError(error['email_exists'], 2)     
		except ValidationError as e:
			messages.add_message(request, messages.ERROR, e.messages[-1]) 
			redirect_path = "/"
			query_string = 'rst=%d' % e.code
			redirect_url = format_redirect_url(redirect_path, query_string)
			return HttpResponseRedirect(redirect_url)

		if not error:
			user.is_active = True
			user.username=request.POST['username']
			print 'username', user.username
			user.email=request.POST['email_id']
			print 'email', user.email
			user.password=request.POST['password']
			print 'pswd', user.password
			user.set_password(user.password)
			user.save()
			print "user saved"
			userprofile = Userprofile()
			userprofile.user_id=user.id
			userprofile.mobile=request.POST['mobile'] 
			userprofile.save() 
			print "userprofile saved"
			send_templated_mail(
			  template_name = 'welcome',
			  subject = 'Welcome Evewat',
			  from_email = 'eventswat@gmail.com',
			  recipient_list = [user.email],
			  context={
					   'user': user.username,
			  },
			)              
			registered = True
			user = User.objects.get(email=user.email)
			print 'user after reg', user
			user.backend='django.contrib.auth.backends.ModelBackend'
			login(request, user)    
			return HttpResponseRedirect('/start/?user_id=' + str(user.id))
	elif user.id is None:
		return HttpResponseRedirect('/')
	else:    
		user_id = user.id
		return render_to_response('index_v2.html', {'user_id':user_id} ,context_instance=RequestContext(request))

# @csrf_exempt
# def register(request): 
#     print "register" 
#     context = RequestContext(request) 
#     registered = False
#     user=User()
#     userprofile=Userprofile()
#     if request.method == 'POST': 
#         email=request.POST['email_id']      
#         username=request.POST['username']              
#         if User.objects.filter(email=email).exists():
#             error={}
#             print 'User.objects.filter(email=email).exists()', User.objects.filter(email=email).exists()
#             error['email_exists'] = True
#             return HttpResponse(json.dumps(error, ensure_ascii=False),mimetype='application/json')
#         if not error:
#             user.is_active = True
#             user.username=request.POST['username']
#             print 'username', user.username
#             user.email=request.POST['email_id']
#             print 'email', user.email
#             user.password=request.POST['password']
#             print 'pswd', user.password
#             user.set_password(user.password)
#             user.save()
#             print "user saved"
#             userprofile = Userprofile()
#             userprofile.user_id=user.id
#             userprofile.mobile=request.POST['mobile'] 
#             userprofile.save() 
#             print "userprofile saved"
#             send_templated_mail(
#               template_name = 'welcome',
#               subject = 'Welcome Evewat',
#               from_email = 'eventswat@gmail.com',
#               recipient_list = [user.email],
#               context={
#                        'user': user.username,
#               },
#             )              
#             registered = True
#             user = User.objects.get(email=user.email)
#             print 'user after reg', user
#             user.backend='django.contrib.auth.backends.ModelBackend'
#             login(request, user)    
#             return HttpResponseRedirect('/start/?user_id=' + str(user.id))
#     elif user.id is None:
#         return HttpResponseRedirect('/')
#     else:    
#         user_id = user.id
#         return render_to_response('index_v2.html', {'user_id':user_id} ,context_instance=RequestContext(request))

@csrf_exempt
@login_required(login_url='/?lst=1')
def post_event(request):
	try:
		category= Category.objects.all()
		subcategory = SubCategory.objects.all()
		state=City.objects.values_list('state',flat=True)
		premium=PremiumPriceInfo.objects.all()
		return render_to_response("post_event.html",{'premium':premium,'subcategory':subcategory,'category':category,'state':list(set(state))}, context_instance=RequestContext(request))
	except:
		return render_to_response("post_event.html",{'message':'Sorry for inconvenience.Some thing went to wrong'}, context_instance=RequestContext(request))
		
def submit_event_v2(request):
	try:
		if request.method=="POST":
			postevent=Postevent()
			postevent.name=request.POST['name']
			postevent.email=request.POST['email']
			postevent.mobile=request.POST.get('mobile','0')
			postevent.event_title=request.POST.get('eventtitle','')
			startdate=request.POST.get('startdate','')
			date,month,year=startdate.split('-')
			postevent.startdate=year+'-'+month+'-'+date
			enddate=request.POST.get('enddate','')
			date,month,year=enddate.split('-')
			postevent.enddate=year+'-'+month+'-'+date
			postevent_category=Category.objects.get(id=request.POST.get('category',''))
			postevent.category=postevent_category
			postevent_subcategory=SubCategory.objects.get(id=request.POST.get('eventtype',''))
			postevent.eventtype=postevent_subcategory
			postevent.eventdescription=request.POST.get('eventdescription','')
			postevent.address=request.POST.get('address','')
			postevent.state=request.POST.get('state','')
			postevent_city=request.POST.get('city','')
			postevent.city=postevent_city
			postevent_college=request.POST.get('college','')
			postevent.college=postevent_college
			postevent.department=request.POST.get('dept','')
			postevent_poster=request.FILES.getlist('poster[]')
		
			def handle_uploaded_file(f):            
				postevent_poster = open(settings.MEDIA_ROOT+'/events/' + '%s' % f.name, 'wb+')
				for chunk in f.chunks():
					postevent_poster.write(chunk)
				postevent_poster.close()
			photosgroup = ''         
			count=len(postevent_poster)
		

			if count :
				for uploaded_file in postevent_poster:
					count=count-1
					handle_uploaded_file(uploaded_file)
					if count==0:
						photosgroup=photosgroup  + '/events/' + str(uploaded_file)
					else:
						photosgroup=photosgroup  + '/events/' +str(uploaded_file) + ','
				postevent.poster=photosgroup
			else:
				postevent.poster='/events/img/logo_150.png'
			# if request.POST.get('plan')!='0':
			#     postevent.payment=request.POST.get('plan')
			postevent.save()
			organizer=Organizer()
			post=Postevent.objects.order_by('-pk')[0]
			organizer.postevent=postevent
			organizer.organizer_name=request.POST.get('organizer_name','')
			organizer.organizer_mobile=request.POST.getlist('organizer_mobile[]','')             
			mobile_number= ''       
			count=len(organizer.organizer_mobile)
			if count:
				for number in organizer.organizer_mobile:
					count=count-1
					if count==0:
						mobile_number= mobile_number + number
					else:
						mobile_number= mobile_number + number + ','                
			organizer.organizer_mobile = mobile_number                
			organizer.organizer_email=request.POST.get('organizer_email','')
			organizer.save()
			send_templated_mail(
				  template_name = 'post_event',
				  subject = 'Post Event',
				  from_email = 'eventswat@gmail.com',
				  recipient_list = [postevent.email],
				  context={

						   'user': postevent.name,
								   
					 },
				   )  
			message="Your data succesfully submitted"

		
		# user_amount=request.POST.get('plan')
		# if user_amount!='0':
		#     return HttpResponseRedirect('/payment_event/')
		# elif user_amount=='0':
		#     response=render_to_response("post_event.html",{'message':message}, context_instance=RequestContext(request))
		#else:
			response= render_to_response("post_event.html",{'message':message}, context_instance=RequestContext(request))
			response.delete_cookie('eventtitle')
			response.delete_cookie('startdate')
			response.delete_cookie('enddate')
		#response.delete_cookie('plan')
			response.delete_cookie('category_name')
			response.delete_cookie('eventtype_name')
			response.delete_cookie('eventtype')
			response.delete_cookie('category')
			response.delete_cookie('eventdescription')
			return response
		else:
			return render_to_response("post_event.html",{'message':'Insufficient data'}, context_instance=RequestContext(request))
	except:
		response = render_to_response("post_event.html",{'message':'Something went to wrong'}, context_instance=RequestContext(request))
		response.delete_cookie('eventtitle')
		response.delete_cookie('startdate')
		response.delete_cookie('enddate')
		response.delete_cookie('plan')
		response.delete_cookie('category_name')
		response.delete_cookie('eventtype_name')
		response.delete_cookie('eventtype')
		response.delete_cookie('category')
		response.delete_cookie('eventdescription')
		return response

@csrf_exempt
@login_required(login_url='/?lst=2')
def upload_banner(request):
	if request.POST.get('price',False):
		uploadbanner=SiteBanner()
		uploadbanner.price=request.POST.get('price',request.COOKIES.get('price'))
		uploadbanner.position=request.POST.get('position',request.COOKIES.get('position'))
		uploadbanner.pageurl=request.POST.get('pageurl',request.COOKIES.get('pageurl'))
		uploadbanner.banner=request.FILES.get('banner',request.COOKIES.get('banner'))
		uploadbanner.link=request.POST['link']
		uploadbanner.save()
		send_templated_mail(
			  template_name = 'banner',
			  subject = 'Uplaod Banner',
			  from_email = 'eventswat@gmail.com',
			  recipient_list = [request.user.email ],
			  context={
					   'user': request.user,
										   
			  },
		  )  
		response=HttpResponseRedirect("/payment/")
		response.set_cookie( 'price', uploadbanner.price )
		response.set_cookie( 'position', uploadbanner.position )
		response.set_cookie( 'banner', uploadbanner.banner )
		response.set_cookie( 'pageurl', uploadbanner.pageurl )
	#field9 is payu success variable
	# this if condition for after success of payment
	elif 'field9' in request.POST:
		message="Your data succesfully uploaded"
		response = render_to_response("uploadbanner.html",{'message':message},context_instance=RequestContext(request))
	else:
		message="Something went to wrong"
		response = render_to_response("uploadbanner.html",{'message':message},context_instance=RequestContext(request))   
	return response

@csrf_exempt
def success(request):
	#field9 is payu success variable
	# this if condition for after success of payment
	if 'field9' in request.POST:
		response = render_to_response("success.html",context_instance=RequestContext(request))
		response.delete_cookie('eventtitle')
		response.delete_cookie('startdate')
		response.delete_cookie('enddate')
		response.delete_cookie('plan')
		response.delete_cookie('category_name')
		response.delete_cookie('eventtype_name')
		response.delete_cookie('eventtype')
		response.delete_cookie('category')
		response.delete_cookie('eventdescription')
	else:
		response =HttpResponseRedirect('/') 
	return response
  

def importcollegedata(request):

  saved = False
  saved_leads = []
  
  if request.method == 'POST':
	form = UploadFileForm(request.POST, request.FILES)
	
	if form.is_valid():
	  
	  f = request.FILES['file']
	  path = os.path.join(settings.MEDIA_ROOT, 'imports')
	  
	  if not os.path.isdir(path):
		os.mkdir(path)
		
	  path = os.path.join(path, "%s_%s"%(request.user.pk, time.time())) 

	  destination = open(path, 'wb+')
	  for chunk in f.chunks():
		destination.write(chunk)
	  destination.close()      
	  
	  
	  workbook = openpyxl.load_workbook(filename=path)      
	  sheet_names = workbook.get_sheet_names()

	  for sheet_name in sheet_names:
		sheet = workbook.get_sheet_by_name(sheet_name)
		if sheet_name == 'Sheet1':            
			rows = sheet.get_highest_row()
			cols = sheet.get_highest_column()

			city_mapping = {}
			collegeaddress_mapping = {}
			collegecategory_mapping = {}
			
			city_fields = [f.attname for f in City._meta.local_fields]
			collegename_fields = [f.attname for f in College._meta.local_fields]
			collegecategory_fields = [f.attname for f in Category._meta.local_fields]
			
			
			if rows > 1 and cols > 0 and sheet.cell(row=0, column=0) is not None:                            
			  for i in range( cols ):              
				v = sheet.cell(row=0, column=i).value
				v = v.replace(' ', '_').lower()                
				
				if v.startswith('city'):
				   v = v.replace('city_', '')
				   if v in city_fields:
					 city_mapping[v] = i
				
				if v.startswith('state'):                   
				   v = v.replace('state_', '')
				   if v in city_fields:
					 city_mapping[v] = i
				
				if v.startswith('country_code'):                   
				   v = v.replace('country_code_', '')
				   if v in city_fields:
					 city_mapping[v] = i
				
				if v.startswith('country_name'):                   
				   v = v.replace('country_name_', '')
				   if v in city_fields:
					 city_mapping[v] = i               
				
				if v.startswith('college_name'):                   
				   v = v.replace('college_name_', '')
				   if v in collegename_fields:
					 collegeaddress_mapping[v] = i

				if v.startswith('name'):                   
				   v = v.replace('name_', '')
				   if v in collegecategory_fields:
					 collegecategory_mapping[v] = i     
				
						
				else:
					pass
				  # if v in lead_fields:
				  #   lead_mapping[v] = i              
			  
			  for i in range(1, rows):                
				citylist = City() if len(city_mapping.keys()) > 0 else None                 
				collegelist = College() if len(collegeaddress_mapping.keys()) > 0 else None
				collegecategorylist = Category() if len(collegecategory_mapping.keys()) > 0 else None                      

				for field in city_mapping:
				  col = city_mapping[field]
				  v = sheet.cell(row=i, column=col)                                
				  if v:                    
					setattr(citylist, field, v.value)

				for field in collegeaddress_mapping:
				  col = collegeaddress_mapping[field]
				  v = sheet.cell(row=i, column=col)                                
				  if v:
					setattr(collegelist, field, v.value)    
				
				for field in collegecategory_mapping:
				  col = collegecategory_mapping[field]
				  v = sheet.cell(row=i, column=col)                                
				  if v:                    
					setattr(collegecategorylist, field, v.value)                
				
				if citylist:
				  if not City.objects.filter(city=citylist.city).exists():
					citylist.save() 

				if collegecategorylist:
				  if not Category.objects.filter(name=collegecategorylist.name).exists():
					collegecategorylist.save()      

				if collegelist:
				  if City.objects.filter(city=citylist.city).exists():
					cityvalue = City.objects.filter(city=citylist.city)                    
					for c in cityvalue:
						collegelist.city_id = c.id
					collegelist.save()
				  
				  if Category.objects.filter(name=collegecategorylist.name).exists():
					categoryvalue = Category.objects.filter(name=collegecategorylist.name)                    
					for c in categoryvalue:
						collegelist.collegetype_id = c.id
					collegelist.save()  
		  
	  saved = True
	  success_import= 'Successfully imported'
	  return render_to_response('import.html', {'success_import':success_import, 'form': form, 'saved':saved, 'saved_leads': saved_leads}, context_instance=RequestContext(request))
  else:
	form = UploadFileForm()
	  
  return render_to_response('import.html', {'form': form, 'saved':saved, 'saved_leads': saved_leads}, 
	context_instance=RequestContext(request))
  
@csrf_exempt
def feedback(request):
	if request.is_ajax():
		if request.method=="POST":
			print "enter in post"
			feedback=Feedback()
			feedback.name=request.POST.get('name')
			print feedback.name
			feedback.email=request.POST.get('email')
			print feedback.email
			feedback.comments=request.POST.get('comments')
			print feedback.comments
			feedback.rating=request.POST.get('rating')
			feedback.save()
			msg = "The operation has been received correctly."
	else:
		msg = "Fail"
	return HttpResponse(msg)

def getstate(request):
	from collections import OrderedDict
	results = []
	unsort_dict = {}
	key_loc = request.GET.get('term')
	state_lists = City.objects.filter(state__icontains=key_loc)

	for state_list in state_lists:
		statename = state_list.state.strip()
		stateid = state_list.id
		unsort_dict[statename] = {'stateid':stateid, 'label':statename, 'value':statename}

	sorted_dic = OrderedDict(sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
	for k, v in sorted_dic.iteritems():  
		results.append(v)

	return HttpResponse(simplejson.dumps(results), mimetype='application/json')

def getcollege(request):
	from collections import OrderedDict
	results = []
	unsort_dict = {}
	key_loc = request.GET.get('term')
	city=request.GET.get('city')
	filterargs = { 'city_id': city, 'college_name__icontains': key_loc }
	college_lists = College.objects.filter(**filterargs)

	for college_list in college_lists:
		collegename = college_list.college_name.strip()
		collegeid = college_list.id
		unsort_dict[collegename] = {'collegeid':collegeid, 'label':collegename, 'value':collegename}

	sorted_dic = OrderedDict(sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
	for k, v in sorted_dic.iteritems():  
		results.append(v)

	return HttpResponse(simplejson.dumps(results), mimetype='application/json')

def getdept(request):
	from collections import OrderedDict
	results = []
	unsort_dict = {}
	key_loc = request.GET.get('term')
	college=request.GET.get('college')
	filterargs={'college__id':college}
	#print CollegeDepartment.objects.filter(college__id=college)
	department_lists = Department.objects.filter(department__icontains=key_loc)
	for department_list in department_lists:
		departmentname = department_list.department.strip()
		departmentid = department_list.id
		unsort_dict[departmentname] = {'departmentid':departmentid, 'label':departmentname, 'value':departmentname}

	sorted_dic = OrderedDict(sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
	for k, v in sorted_dic.iteritems():  
		results.append(v)

	return HttpResponse(simplejson.dumps(results), mimetype='application/json')


def getcity_base(request):
	from collections import OrderedDict
	results = []
	unsort_dict = {}
	key_loc = request.GET.get('term')
	city_lists = City.objects.filter(city__icontains= key_loc)

	for city_list in city_lists:
		cityname = city_list.city.strip()
		cityid = city_list.id
		unsort_dict[cityname] = {'cityid':cityid, 'label':cityname, 'value':cityname}

	sorted_dic = OrderedDict(sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
	for k, v in sorted_dic.iteritems():  
		results.append(v)

	return HttpResponse(simplejson.dumps(results), mimetype='application/json')
	
def event_for_subcategory(request):
	if request.is_ajax() and request.GET and 'sub_category_id' in request.GET:
		objs1 = Postevent.objects.filter(festtype_id=sub_category_id)
		for obj in objs1:
			print obj.brand_name        
		return JSONResponse([{'id': o1.id, 'name': smart_unicode(o1.brand_name)}
			for o1 in objs1])
	else:
		return JSONResponse({'error': 'Not Ajax or no GET'})


# To Load Categories in Home Page as Left Side bar
def all_subcategory_for_category(request):
	if request.is_ajax():
		objs1 = SubCategory.objects.all()
		return JSONResponse([{'name': o1.name, 'id': o1.id, 'icon':smart_unicode(o1.icon), 'hover_icon':smart_unicode(o1.hover_icon)} for o1 in objs1])       
	else:
		return JSONResponse({'error': 'Not Ajax or no GET'})

# To Load SubCategories in Home Page as Left Side bar
def subcategory_for_category(request):
	if request.is_ajax() and request.GET and 'category_id' in request.GET:
		objs1 = SubCategory.objects.filter(category__id=request.GET['category_id']) 
		return JSONResponse([{'name': o1.name, 'id': o1.id, 'icon':smart_unicode(o1.icon), 'hover_icon':smart_unicode(o1.hover_icon)} for o1 in objs1])       
	else:
		return JSONResponse({'error': 'No Ajax or No Get Request'})

def find_colleges(request):
	from django.utils.encoding import smart_unicode, force_unicode
	if request.is_ajax() and request.GET and 'city_id' in request.GET:
		objs = College.objects.filter(city=request.GET['city_id'])
		return JSONResponse([{'id': o.id, 'name': smart_unicode(o.college_name)}
		for o in objs])
	else:
		return JSONResponse({'error': 'Not Ajax or no GET'})

def find_department(request):
	from django.utils.encoding import smart_unicode, force_unicode
	if request.is_ajax() and request.GET and 'college_id' in request.GET:
		objs = CollegeDepartment.objects.filter(college__id=request.GET['college_id'])
		return JSONResponse([{'id': o.department.id, 'name': smart_unicode(o.department)}
		for o in objs])
	else:
		return JSONResponse({'error': 'Not Ajax or no GET'})
	

def find_subcategory(request):
	from django.utils.encoding import smart_unicode, force_unicode
	if request.is_ajax() and request.GET and 'category_id' in request.GET:
		objs = SubCategory.objects.filter(category__id=request.GET['category_id'])
		return JSONResponse([{'id': o.id, 'name': smart_unicode(o.name)}
		for o in objs])
	else:
		return JSONResponse({'error': 'Not Ajax or no GET'})

def find_city(request):
	from django.utils.encoding import smart_unicode, force_unicode
	if request.is_ajax() and request.GET and 'state' in request.GET:
		objs = City.objects.filter(state=request.GET['state'])
		return JSONResponse([{'id': o.id, 'name': smart_unicode(o.city)}
		for o in objs])
	else:
		return JSONResponse({'error': 'Not Ajax or no GET'})

def getcity(request):
	from collections import OrderedDict
	results = []
	unsort_dict = {}
	key_loc = request.GET.get('term')
	state=request.GET.get('state')
	filterargs = { 'state': state, 'city__icontains': key_loc }
	city_lists = City.objects.filter(**filterargs)
	for city_list in city_lists:
		cityname = city_list.city.strip()
		cityid = city_list.id
		unsort_dict[cityname] = {'cityid':cityid, 'label':cityname, 'value':cityname}

	sorted_dic = OrderedDict(sorted(unsort_dict.iteritems(), key=lambda v: v[0]))
	for k, v in sorted_dic.iteritems():  
		results.append(v)

	return HttpResponse(simplejson.dumps(results), mimetype='application/json')

def home_v2(request):
   context = RequestContext(request,
						   {'request': request,
							'user': request.user})
   return render_to_response('home_v2.html',
							 context_instance=context)

def get_events_for_calendar(request):
    import datetime
    events = Postevent.objects.all()
    time = datetime.time(10, 25)
    events_list = []
    for event in events:
        event_data = {'id':str(event.id), 'title':event.event_title, 'start':smart_unicode(datetime.datetime.combine(event.startdate,time)),'end':smart_unicode(datetime.datetime.combine(event.enddate,time))}
        events_list.append(event_data)
    # print "event_list", events_list
    return HttpResponse(simplejson.dumps(events_list), mimetype='application/json')

@csrf_exempt
def user_profile(request):
	if request.user.is_authenticated:
		requested_user=User.objects.get(email=request.user.email)
		if request.method == 'POST':
			last_name=request.POST.get('last_name', '')			
			user_mobile=request.POST.get('mobile', '')
			print 'user_mobile', user_mobile
			gender=request.POST.get('gender', '')		
			Date_of_birth=request.POST.get('dob', '')
			if Date_of_birth=='':
				Date_of_birth=None
			user_address=request.POST.get('address', '')			
			profile_picture = request.FILES.get('profile_poster')
			print 'profile_picture', profile_picture
			def handle_uploaded_file(f):
				print "settings.MEDIA_ROOT", settings.MEDIA_ROOT
				profile_picture = open(settings.MEDIA_ROOT+'/events/' + '%s' % f.name, 'wb+')
				for chunk in f.chunks():
					profile_picture.write(chunk)
				profile_picture.close()
			handle_uploaded_file(profile_picture)
			profile_picture = '/events/' + str(profile_picture)					
			if Userprofile.objects.filter(user_id=requested_user.id).exists():
				user_id=Userprofile.objects.get(user_id=requested_user.id)
				user_id.lastname=last_name
				user_id.mobile=user_mobile
				user_id.gender=gender
				user_id.Date_of_birth=Date_of_birth
				user_id.user_address=user_address
				user_id.profile_pic=profile_picture		
				user_id.save()
				
			else:
				userprofile=Userprofile()
				userprofile.user_id=requested_user.id
				userprofile.lastname=last_name	
				userprofile.mobile=user_mobile
				userprofile.gender=gender
				userprofile.Date_of_birth=Date_of_birth
				userprofile.user_address=user_address
				userprofile.profile_pic=profile_picture
				userprofile.save()			
		try:			     
			requested_user_profile=Userprofile.objects.get(user_id=requested_user.id)		
			events_for_user=Postevent.objects.filter(email=request.user.email)
			print 'events_for_user', events_for_user			
			return render_to_response("user_profile.html", {'requested_user':requested_user, 'requested_user_profile':requested_user_profile, 'events_for_user':events_for_user}, context_instance=RequestContext(request))        
		except:			
			events_for_user=Postevent.objects.filter(email=request.user.email)
			return render_to_response("user_profile.html", {'requested_user':requested_user, 'events_for_user':events_for_user}, context_instance=RequestContext(request))
@csrf_exempt
def privacy(request):
	print 'request.user.email',request.user.email
	u = User.objects.get(email=request.user.email)
	new_password=request.POST.get('newpassword')
	u.set_password(new_password)
	u.save()
	return render_to_response("user_profile.html", context_instance=RequestContext(request))




		
